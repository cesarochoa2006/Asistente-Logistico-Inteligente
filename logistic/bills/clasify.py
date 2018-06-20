import openpyxl
import numpy as np
from sklearn import tree,svm
import random as ran
"""
class Repartidor:
    def __init__(self, nombre, envios):
        self.nombre = nombre
        self.envios = envios
class Envio:
    def __init__(self, celda,cliente,factura, valorfact, fechamin, fechamax, fechaentrega, operador, fecharepro,direccion,prioridad,distancia):
        self.celda = celda
        self.cliente=cliente
        self.numfact=factura
        self.valorfact = valorfact
        self.fechamin = fechamin
        self.fechamax = fechamax
        self.fechaentrega = fechaentrega
        self.operador = operador
        self.fecharepro = fecharepro
        self.prioridad=prioridad
        self.distancia=distancia
        self.direccion = direccion

    def to_list(self):
        res = []
        res.append(self.celda)
        res.append(self.cliente)
        res.append(self.numfact)
        res.append(self.valorfact)
        res.append(self.fechamin)
        res.append(self.fechamax)
        res.append(self.fechaentrega)
        res.append(self.operador)
        res.append(self.direccion)
        res.append(self.prioridad)
        res.append(self.distancia)
        return res
"""
def to_numbers(list):
    _, numbers = np.unique(list, return_inverse=True)
    return numbers


def table_head(filepath):
    if filepath:
        doc = openpyxl.load_workbook(filepath)
        hoja1 = doc.get_sheet_by_name('Hoja1')
        doc.close()
        return [
            hoja1.cell(row=1,column=2).value,
            hoja1.cell(row=1, column=3).value,
            hoja1.cell(row=1, column=4).value,
            hoja1.cell(row=1, column=6).value,
            hoja1.cell(row=1, column=7).value,
            hoja1.cell(row=1, column=8).value,
            hoja1.cell(row=1, column=12).value,
            #hoja1.cell(row=1, column=11).value,
            hoja1.cell(row=1, column=10).value,
        ]

def list_of_list(filepath):
    if filepath:
        doc = openpyxl.load_workbook(filepath)
        hoja1 = doc.get_sheet_by_name('Hoja1')

        lista_envios = []
        for i in range(93 - 1):
            cliente = hoja1.cell(row=i + 2, column=2).value  # se guarda la variable cliente
            factura = hoja1.cell(row=i + 2, column=3).value  # se guarda la variable factura
            valorfact = hoja1.cell(row=i + 2, column=4).value  # se guarda la variable valor de factura
            fechamin = hoja1.cell(row=i + 2, column=6).value  # se guarda la fecha minima
            fechamax = hoja1.cell(row=i + 2, column=8).value  # se guarda la fecha maxima
            fechaentrega = hoja1.cell(row=i + 2, column=7).value  # se guarda la fecha de entrega
            operador = hoja1.cell(row=i + 2, column=10).value  # se guarda el operador
            fecharepro = hoja1.cell(row=i + 2, column=11).value  # se guarda la reprogramacion
            direccion = hoja1.cell(row=i + 2, column=12).value  # se guarda la direccion
            fechamin = list(map(int, fechamin.strip().split("-")))
            fechamax = list(map(int, fechamax.strip().split("-")))
            fechaentrega = list(map(int, fechaentrega.strip().split("-")))


            row = [cliente, factura, valorfact, fechamin, fechamax, fechaentrega, direccion,operador]
            lista_envios.append(row)

            doc.close()
        return lista_envios
def cla_source_list(filepath):
    if filepath:
        doc = openpyxl.load_workbook(filepath)
        hoja1 = doc.get_sheet_by_name('Hoja1')

        lista_envios = []
        for i in range(93 - 1):
            cliente = hoja1.cell(row=i + 2, column=2).value  # se guarda la variable cliente
            factura = hoja1.cell(row=i + 2, column=3).value  # se guarda la variable factura
            valorfact = hoja1.cell(row=i + 2, column=4).value  # se guarda la variable valor de factura
            fechamin = hoja1.cell(row=i + 2, column=6).value  # se guarda la fecha minima
            fechamax = hoja1.cell(row=i + 2, column=8).value  # se guarda la fecha maxima
            fechaentrega = hoja1.cell(row=i + 2, column=7).value  # se guarda la fecha de entrega
            #operador = hoja1.cell(row=i + 2, column=10).value  # se guarda el operador
            #fecharepro = hoja1.cell(row=i + 2, column=11).value  # se guarda la reprogramacion
            direccion = hoja1.cell(row=i + 2, column=12).value  # se guarda la direccion
            fechamin = list(map(int, fechamin.strip().split("-")))
            fechamax = list(map(int, fechamax.strip().split("-")))
            fechaentrega = list(map(int, fechaentrega.strip().split("-")))


            row = [cliente, factura, valorfact, fechamin, fechamax, fechaentrega, direccion]
            lista_envios.append(row)

            doc.close()
        return lista_envios

def get_pending(filepath):
    """if filepath is None:
        filepath=os.path.join(MEDIA_ROOT,'DATOS.xlsx') or None""" #default filepath
    doc = openpyxl.load_workbook(filepath)
    hoja3 = doc.get_sheet_by_name('rep')
    pending = []
    for i in range(15):
        cliente = hoja3.cell(row=i + 2, column=2).value  # se guarda la variable cliente
        factura = hoja3.cell(row=i + 2, column=3).value  # se guarda la variable factura
        valorfact = hoja3.cell(row=i + 2, column=4).value  # se guarda la variable valor de factura
        fechamin = str(hoja3.cell(row=i + 2, column=6).value)  # se guarda la fecha minima
        fechamax = str(hoja3.cell(row=i + 2, column=7).value)  # se guarda la fecha maxima
        # operador = hoja3.cell(row=i + 2, column=10).value  # se guarda el operador
        # fecharepro = hoja3.cell(row=i + 2, column=11).value  # se guarda la reprogramacion
        direccion = hoja3.cell(row=i + 2, column=12).value  # se guarda la direccion
        fechamin = list(map(int, fechamin.strip().split("-")))
        fechamax = list(map(int, fechamax.strip().split("-")))

        row = [cliente, factura, valorfact, fechamin, fechamax, fechamin, direccion]
        pending.append(row)
    return pending

def data_clasification(filepath):
    """
    if filepath is None:
        filepath=os.path.join(MEDIA_ROOT,'DATOS.xlsx') or None #default filepath
    """
    target = []
    try:
        pending = get_pending(filepath)
        raw_factor=int(len(pending)*0.3)
        #print (raw_factor)
        #print (len(data))

        #target_names = ['SANCHEZ POLO','EXXE','TCC','BLU LOGISTICS']
        #feature_names=['cliente','factura','valor_factura','fechamin','fechamax']


        data = list_of_list(filepath)
        #print (data[0])
        clients = []

        mindate=[]
        maxdate=[]
        progdate=[]
        dir=[]
        for d in data:
            clients.append(d[0])

            target.append(d[7])
            mindate.append(d[3])
            maxdate.append(d[4])
            progdate.append(d[5])
            dir.append(d[6])
        data=cla_source_list(filepath)
        clients= to_numbers(clients)
        target_names, _ = np.unique(target,return_inverse=True)
        #print (target_names)
        mindate=to_numbers(mindate)
        maxdate=to_numbers(maxdate)
        progdate=to_numbers(progdate)
        dir=to_numbers(dir)
        #print (len(clients),len(operators))
        for i in range(len(data)):
            data[i][0]=clients[i]

            data[i][3]=mindate[i]
            data[i][4]=maxdate[i]
            data[i][5]=progdate[i]
            data[i][6]=dir[i]
        #print (data[0])
        clients=[]
        mindate=[]
        maxdate=[]
        dir=[]
        for p in pending:
            clients.append(p[0])
            mindate.append(p[3])
            maxdate.append(p[4])
            dir.append(p[6])
        clients=to_numbers(clients)
        mindate = to_numbers(mindate)
        maxdate = to_numbers(maxdate)
        dir = to_numbers(dir)
        for i in range(len(pending)):
            pending[i][0]=clients[i]
            pending[i][3]=mindate[i]
            pending[i][4]=maxdate[i]
            pending[i][5] = maxdate[i]
            pending[i][6]=dir[i]
        #print (pending[0])
        #clasify operator

        train_target = target
        train_data= data
        #print (target[0],data[0])
        test_data=[pending[0]]
        #clf = tree.DecisionTreeClassifier()
        #clf.fit(train_data,train_target)
        #print(clf.predict(test_data))
        def clasify_data(train_target,train_data,test_data):
            #clf = tree.DecisionTreeClassifier()
            clf = svm.SVC()
            clf.fit(train_data,train_target)
            return clf.predict(test_data)


        clasify_results=[]
        for i in range(len(pending)-raw_factor):
            clasify_results.append(clasify_data(target,data,[pending[i]])),
        for i in range(raw_factor):
            ap = []
            ap.append(target_names[ran.randint(0,len(target_names)-1)])
            clasify_results.append(ap)
        #return_string=[]
        pending = get_pending(filepath)


        for i in range(len(pending)):
            pending[i].append(clasify_results[i][0])
        for row in pending:
            print (row)
        return (pending)

    except:
        print("Filepath is None")
#print(data_clasification(None))

# def Obtenerlista(filepath):
#     doc = openpyxl.load_workbook(filepath)
#     # print(doc.get_sheet_names())
#     hoja1 = doc.get_sheet_by_name('Hoja1')
#     # print(hoja.cell(row=2,column=10).value)
#     texto = ""
#     lista_envios = []
#     for i in range(93 - 1):
#         cliente = hoja1.cell(row=i + 2, column=2).value#se guarda la variable cliente
#         factura = hoja1.cell(row=i + 2, column=3).value#se guarda la variable factura
#         valorfact = hoja1.cell(row=i + 2, column=4).value#se guarda la variable valor de factura
#         fechamin = hoja1.cell(row=i + 2, column=6).value#se guarda la fecha minima
#         fechamax = hoja1.cell(row=i + 2, column=7).value#se guarda la fecha maxima
#         fechaentrega = hoja1.cell(row=i + 2, column=8).value#se guarda la fecha de entrega
#         operador = hoja1.cell(row=i + 2, column=10).value#se guarda el operador
#         fecharepro = hoja1.cell(row=i + 2, column=11).value#se guarda la reprogramacion
#         direccion = hoja1.cell(row=i + 2, column=12).value#se guarda la direccion
#         fechamin = list(map(int, fechamin.strip().split("-")))
#         fechamax = list(map(int, fechamax.strip().split("-")))
#         fechaentrega = list(map(int, fechaentrega.strip().split("-")))
#
#         if(fechaentrega[1] == fechamax[1]):
#           prioridad = fechaentrega[2]-fechamax[2]
#         else:
#           prioridad = fechaentrega[2]+30-fechamax[2]
#
#         envio = Envio(i + 2,cliente, valorfact, fechamin, fechamax, fechaentrega, operador, fecharepro,direccion,prioridad,0)
#         lista_envios.append(envio)
#         # doc.save("DATOS.xlsx")
#     doc.close()
#     return lista_envios
#
# def clasificacion_clase(lista_envios,kvecinos):
#     clases_nombre =[]
#     clases_valor=[]
#     for i in range(kvecinos):
#         if(lista_envios[i].operador in clases_nombre):
#             #print("sumo")
#             clases_valor[clases_nombre.index(lista_envios[i].operador)] = clases_valor[clases_nombre.index(lista_envios[i].operador)] + 1
#         else:
#             #print("creo")
#             clases_nombre.append(lista_envios[i].operador)
#             clases_valor.append(1)
#     #print(clases_nombre)
#     #print(clases_valor)
#     #print("la clase ganadora es: "+str(clases_nombre[clases_valor.index(max(clases_valor))]))
#     #return str(clases_nombre[clases_valor.index(max(clases_valor))])
#
#
# def clasificador(lista_envios,objeto):
#     #test data
#     if len(objeto)<1:
#         celda = 94
#         cliente = "Almacenes Exito S.A."
#         valorfact = "500000"
#         fechamin = [ 17, 9, 5]
#         fechamax = [ 17, 9, 9]
#         fechaentrega = [0,0,0] #dato a clasificar
#         operador = "" #dato a clasificar
#         fecharepro = "NA/A"
#         direccion = "Parque Industrial Malambo PIMSA"
#     else:
#         celda = objeto[0]
#         cliente = objeto[1]
#         valorfact = objeto[2]
#         fechamin = objeto[3]
#         fechamax = objeto[4]
#         fechaentrega = [0, 0, 0]  # dato a clasificar
#         operador = ""  # dato a clasificar
#         fecharepro = objeto[5]
#         direccion = objeto[6]
#     envio_clasificar = Envio(celda,cliente, valorfact , fechamin, fechamax, fechaentrega, operador, fecharepro,direccion,0,0)
#     lista_similares = []
#
#     for i in range(len(lista_envios)):
#
#         if(lista_envios[i].cliente == envio_clasificar.cliente):
#             lista_envios[i].distancia = ((int(envio_clasificar.valorfact)-lista_envios[i].valorfact)**2)**(1/2)
#             lista_similares.append(lista_envios[i])
#
#     lista_similares.sort(key=lambda Envio: Envio.distancia)
#     #mostrarListaEnvios(lista_similares);
#     kvecinos= 5
#     prioridad = 0
#     #print(mostrarListaEnvios(lista_similares))
#     for i in range(kvecinos):
#         prioridad=prioridad +lista_similares[i].prioridad
#
#     prioridad = prioridad/kvecinos
#     operador=clasificacion_clase(lista_similares,kvecinos)
#     #print(prioridad)
#
#     envio_clasificar.fechaentrega = [fechamax[0],fechamax[1],fechamax[2]+int(prioridad)]
#     envio_clasificar.operador = operador
#     envio_clasificar.prioridad = int(prioridad)
#     #print(mostrarEnvio(envio_clasificar))
#
# #clasificador(Obtenerlista(),objetoaclasificar)